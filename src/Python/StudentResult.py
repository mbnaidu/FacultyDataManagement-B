import sys, json, numpy as np
import openpyxl
import json

#Read data from stdin
def read_in():
    lines = sys.stdin.readlines()
    #Since our input would only be having one line, parse our JSON data from that
    return lines
def set_marks(credit):
    if(credit == 'O'):
        return 10
    if(credit == 'S'):
        return 9
    if(credit == 'A'):
        return 8
    if(credit == 'B'):
        return 7
    if(credit == 'C'):
        return 6
    if (credit == 'D'):
        return 5
def main():
    excelPath = "./src/Files/newOutput.xlsx"
    wb=openpyxl.load_workbook(excelPath)
    ws = wb.active
    rows = ws.max_row
    lines = read_in()
    arr = []
    arr = lines[0]
    arr = json.loads(arr)
    presentSem = []
    presentSem = arr['presentSem']
    maleBacklogs = arr['maleBacklogs']
    femaleBacklogs = arr['femaleBacklogs']
    for i in range(len(arr['students'])):
        temp = arr['students'][i]
        output = []
        count = 0
        creditsObtained = 0
        creditsTotal = 0
        for j in range(1, rows + 1):
            if(temp['registerID'] == ws.cell(j, 1).value):
                for k in range(len(temp['semesters'])):
                    if(temp['semesters'][k]['name'] == sys.argv[1]):
                        if ws.cell(j, 5).value == "0":
                            count = count + 1
                        if ws.cell(j, 5).value == "1":
                            credit1 = set_marks(ws.cell(j, 4).value) * 1
                            creditsObtained = creditsObtained + credit1
                            creditsTotal = creditsTotal + 1
                        if ws.cell(j, 5).value == "2":
                            credit2 = set_marks(ws.cell(j, 4).value) * 2
                            creditsObtained = creditsObtained + credit2
                            creditsTotal = creditsTotal + 2
                        if ws.cell(j, 5).value == "3":
                            credit3 = set_marks(ws.cell(j, 4).value) * 3
                            creditsObtained = creditsObtained + credit3
                            creditsTotal = creditsTotal + 3
                        output.append({'subcode': ws.cell(j, 2).value,'subname': ws.cell(j, 3).value,'grade': ws.cell(j, 4).value,'credits': ws.cell(j, 5).value, 'noOfAttempts': 0})
                        temp['semesters'][k][sys.argv[1]] = output
                        temp['semesters'][k]['isAvailable'] = True
                        temp['semesters'][k]['TC'] = creditsTotal
                        temp['semesters'][k]['OC'] = creditsObtained
                        temp['backlogs'] = count
        if count > 0 and temp['gender'] == ('M' or 'Male' or 'm' or 'male' or 'MALE'):
            maleBacklogs = maleBacklogs + 1
        if count > 0 and temp['gender'] == ('F' or 'Female' or 'f' or 'female' or 'FEMALE'):
            femaleBacklogs = femaleBacklogs + 1
    presentSem.append({'name': sys.argv[1], 'noOfAttempts': 0})
    arr['maleBacklogs'] = maleBacklogs
    arr['femaleBacklogs'] = femaleBacklogs
    arr['presentSem'] = presentSem
    arr = json.dumps(arr)
    print(arr)
#start process
if __name__ == '__main__':
    main()