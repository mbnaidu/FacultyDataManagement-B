import sys, json, numpy as np
import openpyxl
import json

#Read data from stdin
def read_in():
    lines = sys.stdin.readlines()
    #Since our input would only be having one line, parse our JSON data from that
    return lines

def main():
    excelPath = "./Assets/newOutput.xlsx"
    wb=openpyxl.load_workbook(excelPath)
    ws = wb.active
    rows = ws.max_row
    lines = read_in()
    checking = []
    arr = []
    arr = lines[0]
    arr = json.loads(arr)
    presentSem = []
    presentSem = arr['presentSem']
    maleBacklogs = arr['maleBacklogs']
    femaleBacklogs = arr['femaleBacklogs']
    for i in range(1, rows + 1):
        count = 0
        if(ws.cell(i, 5).value  != "No Change"):
            for j in range(len(arr['students'])):
                temp = arr['students'][j]
                backlog = temp['backlogs']
                if(temp['registerID'] == ws.cell(i, 1).value):
                    for k in range(len(temp['semesters'])):
                        if(temp['semesters'][k]['name'] == sys.argv[1]):
                            for l in range(len(temp['semesters'][k][sys.argv[1]])):
                                if(temp['semesters'][k][sys.argv[1]][l]['subcode'] == ws.cell(i, 2).value):
                                    count = count + 1
                                    temp['semesters'][k][sys.argv[1]][l]['grade'] = ws.cell(i, 4).value
                                    temp['semesters'][k][sys.argv[1]][l]['credits'] = ws.cell(i, 5).value
                                    temp['backlogs'] = backlog - 1
                                    if count > 0 and temp['gender'] == ('M' or 'Male' or 'm' or 'male' or 'MALE'):
                                        maleBacklogs = maleBacklogs - 1
                                    if count > 0 and temp['gender'] == ('F' or 'Female' or 'f' or 'female' or 'FEMALE'):
                                        femaleBacklogs = femaleBacklogs - 1
    arr['maleBacklogs'] = maleBacklogs
    arr['femaleBacklogs'] = femaleBacklogs
    arr = json.dumps(arr)
    print(arr)
#start process
if __name__ == '__main__':
    main()