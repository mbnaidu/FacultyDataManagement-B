import sys
import tabula
import os
from os.path import exists
from PyPDF2 import PdfFileWriter, PdfFileReader
import openpyxl
import json

is_resultPDF_path = './Assets/results.pdf'
is_studentsListDatapath = "./Assets/studentsList.xlsx"
is_resultPDF = exists(is_resultPDF_path)
is_studentsListData = exists(is_studentsListDatapath)

if(is_resultPDF):
    required_pages = sys.argv[1].split(',')
    for i in range(0, len(required_pages)):
        required_pages[i] = int(required_pages[i])
    pages_to_add = required_pages # page numbering starts from 0
    pdfFileObj = open(is_resultPDF_path, 'rb')
    infile = PdfFileReader(pdfFileObj, strict=False)
    output = PdfFileWriter()
    for i in range(infile.getNumPages()):
        if i in pages_to_add:
            p = infile.getPage(i)
            output.addPage(p)
    with open('./Assets/removed.pdf', 'wb') as f:
        output.write(f)
    file_exists = exists('./Assets/removed.pdf')
    if(file_exists):
        tables = tabula.read_pdf("./Assets/removed.pdf", pages="all")
        tabula.convert_into("./Assets/removed.pdf", "./Assets/output.xlsx", output_format="csv", pages="all")
else:
    if(is_studentsListData):
        arr = []
        wb=openpyxl.load_workbook(is_studentsListDatapath)
        ws = wb.active
        rows = ws.max_row
        for i in range(2, rows):
            arr.append({'registerID': ws.cell(i, 1).value, 'fullName': ws.cell(i, 2).value, 'backlogs': 0, 'percentage': 0, '1-1':[], '1-2':[], '2-1':[], '2-2':[], '3-1':[], '3-2':[], '4-1':[], '4-2':[]})
        arr = json.dumps(arr)
        print(arr)
    else:
        print("error reading file")