import os
from os.path import exists
import openpyxl
import json

is_studentsListDatapath = "./src/Files/studentsList.xlsx"
is_studentsListData = exists(is_studentsListDatapath)

if(is_studentsListData):
    arr = []
    wb=openpyxl.load_workbook(is_studentsListDatapath)
    ws = wb.active
    rows = ws.max_row
    for i in range(2, rows + 1):
        arr.append({
            'registerID': ws.cell(i, 1).value, 
            'fullName': ws.cell(i, 2).value, 
            'gender': ws.cell(i, 3).value,
            'backlogs': 0,
            'percentage': 0,
            'semesters':[
                {'sem1Data':[], 'isAvailable': False, 'name':'sem1Data', 'TC': 0, 'OC': 0},
                {'sem2Data':[], 'isAvailable': False, 'name':'sem2Data', 'TC': 0, 'OC': 0},
                {'sem3Data':[], 'isAvailable': False, 'name':'sem3Data', 'TC': 0, 'OC': 0},
                {'sem4Data':[], 'isAvailable': False, 'name':'sem4Data', 'TC': 0, 'OC': 0},
                {'sem5Data':[], 'isAvailable': False, 'name':'sem5Data', 'TC': 0, 'OC': 0},
                {'sem6Data':[], 'isAvailable': False, 'name':'sem6Data', 'TC': 0, 'OC': 0},
                {'sem7Data':[], 'isAvailable': False, 'name':'sem7Data', 'TC': 0, 'OC': 0},
                {'sem8Data':[], 'isAvailable': False, 'name':'sem8Data', 'TC': 0, 'OC': 0},
            ],
            'placements':[]
        })
    arr = json.dumps(arr)
    print(arr)
else:
    print("error reading file")