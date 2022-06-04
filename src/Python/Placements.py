import sys, json, numpy as np
import openpyxl
import json

#Read data from stdin
def read_in():
    lines = sys.stdin.readlines()
    #Since our input would only be having one line, parse our JSON data from that
    return lines

def main():
    excelPath = "./src/Files/studentsList.xlsx"
    wb=openpyxl.load_workbook(excelPath)
    ws = wb.active
    rows = ws.max_row
    lines = read_in()
    arr = []
    checking = []
    arr = lines[0]
    arr = json.loads(arr)
    for i in range(len(arr['students'])):
        temp = arr['students'][i]
        placements = []
        placements = temp['placements']
        for j in range(1, rows + 1):
            if(temp['registerID'] == ws.cell(j, 2).value):
                placements.append({'companyName' : ws.cell(j, 4).value, 'package':  ws.cell(j, 5).value, 'date': str(ws.cell(j, 6).value)})
        temp['placements'] = placements
    arr = json.dumps(arr)
    print(arr)
#start process
if __name__ == '__main__':
    main()